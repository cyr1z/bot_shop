import random
from mongoengine import connect
from models.cats_and_products import Category, Product , Texts
from openpyxl import load_workbook
from faker import Faker
from mtranslate import translate
import imgspy

wb = load_workbook(filename='data.xlsx', read_only=True)
ws = wb['Sheet0']

#
# import datadotworld as dw
# query = "SELECT   name, categories, prices_amountmax, weight, " \
#         "imageurls, brand, primarycategories " \
#         "FROM datafinitielectronicsproductspricingdata " \
#         "GROUP BY name LIMIT 5000"
# data_set = 'datafiniti/electronic-products-and-pricing-data'
#
# results = dw.query(data_set, query)


random_bool = (True, False)
fake = Faker('en_US')
fake_ru = Faker('ru_RU')

if __name__ == "__main__":
    connect("bot_shop")

    Category and subcategory seeder from data.xlsx file

    cats_dict = {}
    a = []
    for row in ws.iter_rows():
        a.append(row[1].value)
    a = list(set(a))

    for i in a:
        cats_dict[i] = {}
        cats_dict[i]['sub_categories'] = []

    for row in ws.iter_rows():
        if row[2].value not in cats_dict[row[1].value]['sub_categories'] \
                and row[2].value != row[1].value:
            cats_dict[row[1].value]['sub_categories'].append(row[2].value)

    for i, v in cats_dict.items():
        print(i, '  >> ')
        subs_list = []
        for sub in cats_dict[i]['sub_categories']:
            print(sub)
            sub_cat = Category(title=sub,
                               title_ru=translate(sub, "ru", "en"),
                               description=fake.text(max_nb_chars=200),
                               description_ru=fake_ru.text(max_nb_chars=200))
            sub_cat.sub_categories = []
            sub_cat.save()
            subs_list.append(sub_cat)
        cat = Category(title=i,
                       title_ru=translate(i, "ru", "en"),
                       description=fake.text(max_nb_chars=200),
                       description_ru=fake_ru.text(max_nb_chars=200))
        cat.sub_categories = subs_list
        cat.save()

    Product seeder from data.xlsx file

    prod_dict = {}
    for row in ws.iter_rows():
        prod_dict[row[0].value] = {}

    for row in ws.iter_rows():
        print(row[0].value)
        prod_dict[row[0].value]['title'] = row[0].value
        prod_dict[row[0].value]['title_ru'] = \
            translate(row[0].value, "ru", "en")
        prod_dict[row[0].value]['description'] = \
            fake.text(max_nb_chars=50)
        prod_dict[row[0].value]['description_ru'] = \
            fake_ru.text(max_nb_chars=50)
        prod_dict[row[0].value]['price'] = \
            int(round(float(row[3].value) * 100, 0))
        prod_dict[row[0].value]['quantity'] = random.randint(2, 100)
        prod_dict[row[0].value]['category'] = \
            Category.objects.filter(title=row[2].value).first()
        print(prod_dict[row[0].value]['title_ru'], end='\n\n')

        weight_calc = row[4].value.split(' ')
        if weight_calc[1] == 'pounds' or weight_calc[1] == 'Pounds':
            prod_dict[row[0].value]['weigth'] = \
                round(float(weight_calc[0]) * 453.592, 0)
        if weight_calc[1] == 'oz' or weight_calc[1] == 'ounces':
            prod_dict[row[0].value]['weigth'] = \
                round(float(weight_calc[0]) * 28.3495, 0)
        if weight_calc[1] == 'lb' or weight_calc[1] == 'lbs':
            prod_dict[row[0].value]['weigth'] = \
                round(float(weight_calc[0]) * 453.592, 0)
        if weight_calc[1] == 'g':
            prod_dict[row[0].value]['weigth'] = round(float(weight_calc[0]), 0)
        if weight_calc[1] == 'kg' or weight_calc[1] == 'Kg':
            prod_dict[row[0].value]['weigth'] = \
                round(float(weight_calc[0]) * 1000, 0)

        photos = str(row[5].value).split(',')
        prod_dict[row[0].value]['photos'] = []
        counter = 0
        for i in photos:
            url, sep, tail = i.partition('%')
            if counter < 10:
                # image url resist and min size test
                try:
                    w = imgspy.info(url)['width']
                    h = imgspy.info(url)['height']
                    print(w, 'x', h)
                    if int(w) >= 300 and int(h) >= 100:
                        prod_dict[row[0].value]['photos'].append(url)
                        counter += 1
                        print('!!!  ', url)
                except:
                    print('bad url', url)

    for i, v in prod_dict.items():
        print(i)
        Product(**v).save()

    text = dict(
        title='Greetings',
        text=random_string(2000)
    )
    Texts(**text).save()
    cats = seed_and_get_categories(10)
    seed_products(50, cats)
